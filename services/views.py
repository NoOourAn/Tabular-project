from django.shortcuts import render
from .models import Timetables
from django.http import HttpResponse
from services.utils import render_to_pdf
import openpyxl
from services import TimeTable
from datetime import datetime, timedelta

to_pdf = []
startdate = None
duedate = None
accesscode = None
# timeslotss = None
# timeslot = None
studentdb = None
subjectdb = None
roomdb = None
wb = None
numberofexams = None
check = None
time_available = []


def home(request):
    return render(request, 'services/home.html')


def step1(request):
    return render(request, 'services/manual-data.html')


def step2(request):
    if request.method == 'POST':
        global startdate, duedate, accesscode
        startdate = request.POST['startdate']
        duedate = request.POST['duedate']
        accesscode = request.POST['accesscode']
    else:
        startdate = None
        duedate = None
        accesscode = None
        return render(request, 'services/manual-data.html')


    print(startdate)
    print(duedate)
    print(accesscode)
    return render(request, 'services/timeslots.html')


def step3(request):
    global numberofexams,startdate,duedate
    if request.method == 'POST':

        numberofexams = request.POST['numberofexams']
        numberofexamsperday = request.POST['numberofexamsperday']

        numberofexams = int(numberofexams)
        numberofexamsperday = int(numberofexamsperday)

        ##########################################

        # startdate = request.POST['startdate']
        # duedate = request.POST['duedate']
        generalexamstarttime = request.POST['generalexamstarttime']
        generalexamendtime = request.POST['generalexamendtime']

        examduration_hrs = request.POST['examdurationH']
        examduration_min = request.POST['examdurationM']

        examduration_hrs = int(examduration_hrs)
        examduration_min = int(examduration_min)

        if examduration_hrs == 0:

            exdurationM_minus_15 = examduration_min - 15
            exdurationM_minus_30 = examduration_min - 30
            exdurationM_minus_45 = examduration_min - 45
            exdurationM_minus_60 = examduration_min - 60

            if exdurationM_minus_15 < 0:
                exdurationM_minus_15 *= -1
            if exdurationM_minus_30 < 0:
                exdurationM_minus_30 *= -1
            if exdurationM_minus_45 < 0:
                exdurationM_minus_45 *= -1
            if exdurationM_minus_60 < 0:
                exdurationM_minus_60 *= -1

            if exdurationM_minus_15 < exdurationM_minus_30 and \
               exdurationM_minus_15 < exdurationM_minus_45 and \
               exdurationM_minus_15 < exdurationM_minus_60:
                examduration_min = 15
                print(15)
            else:
                if exdurationM_minus_30 < exdurationM_minus_45 and \
                   exdurationM_minus_30 < exdurationM_minus_60:
                    examduration_min = 30
                    print(30)
                else:
                    if exdurationM_minus_45 < exdurationM_minus_60:
                        examduration_min = 45
                        print(45)
                    else:
                        examduration_min = 60
                        print(60)


        ##########################################

        startdate = datetime.strptime(startdate, "%Y-%m-%dT%H:%M")
        duedate = datetime.strptime(duedate, "%Y-%m-%dT%H:%M")
        generalexamstarttime = datetime.strptime(generalexamstarttime, "%H:%M")
        generalexamendtime = datetime.strptime(generalexamendtime, "%H:%M")

        print(startdate)
        print(duedate)

        x = 1
        global time_available

        while True:

            laststartdate = startdate
            startdate += timedelta(hours=examduration_hrs, minutes=examduration_min)
            if startdate.hour <= generalexamendtime.hour:
                time_available.append([x,
                                       laststartdate.strftime("%H:%M")
                                       + " - "
                                       + startdate.strftime("%H:%M"),
                                       laststartdate.strftime("%A").upper(),
                                       laststartdate.strftime("%d/%m")
                                       ])
                x += 1

            if startdate.hour >= generalexamendtime.hour:
                # lw al youm antha 5o4 3la alyoum aly wrah
                startdate += timedelta(days=1)
                startdate = datetime(year=startdate.year,
                                                 month=startdate.month,
                                                 day=startdate.day,
                                                 hour=generalexamstarttime.hour,
                                                 minute=generalexamstarttime.minute)

            if startdate >= duedate:
                break

        print(time_available)

    else:
        numberofexams = None
        return render(request, 'services/manual-data.html')

    return render(request, 'services/excel-sheet.html')


# def step3(request, tt_id):
#     timetable = get_object_or_404(Timetables , accessCode = tt_id)
#     return render(request, 'services/generated-timetable.html' ,{'tt':timetable})

def step4(request):

    if request.method == 'POST':
        global studentdb, roomdb, subjectdb, wb
        doc = request.FILES  # returns a dict-like object
        studentdb = openpyxl.load_workbook(doc['st-db'])
        subjectdb = openpyxl.load_workbook(doc['sub-db'])
        roomdb = openpyxl.load_workbook(doc['rm-db'])
    else:
        return render(request, 'services/manual-data.html')

    return render(request, 'services/generated-timetable.html', {'wait': 'We will send email AS soon As your Timetable is READY'})

def fetchTT(request):

    if request.method == 'POST':
        try:
            tables = Timetables.objects.get(org = request.user)
            return render(request, 'services/org-dashboard.html', {'tables': tables})
        except Timetables.DoesNotExist:
            # elif table.DoesNotExist:
            return render(request, 'services/org-dashboard.html', {'error': 'you do not have any timetables yet !'})
    else:
        return render(request, 'services/home.html')


from services import dbconvert

# assign = dbconvert.assign_filename()

def boom(request):

    if startdate is None or numberofexams is None or studentdb is None:
        return render(request, 'services/manual-data.html')

    dbconvert.fetch_data.assign.set_StudentsFile(studentdb)
    dbconvert.fetch_data.assign.set_SubjectsFile(subjectdb)
    dbconvert.fetch_data.assign.set_RoomsFile(roomdb)
    dbconvert.fetch_data.assign.set_TimeSlots(time_available)
    print("zzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzzz")
    print(dbconvert.fetch_data.assign.get_SubjectsFile())
    print(dbconvert.fetch_data.assign.get_StudentsFile())
    print(dbconvert.fetch_data.assign.get_RoomsFile())
    tt = TimeTable.generateTT()
    global to_pdf
    to_pdf = tt
    TT = Timetables()
    TT.startDate = startdate
    TT.dueDate = duedate
    TT.accessCode = accesscode
    TT.org = request.user
    TT.exams = tt
    TT.save()
    timeslots = TimeTable.get_timeslots()
    timeslots.sort
    subjects = {}
    for i in tt:
        if i[4] not in timeslots:
            timeslots.append(i[4])
        if i[3] in subjects:
            templist = subjects[i[3]]
            templist[i[4]] = [i[0], i[1], i[2]]
            subjects[i[3]] = templist
        else:
            templist = {i[4]: [i[0], i[1], i[2]]}
            subjects[i[3]] = templist
    return render(request, 'services/generated-timetable.html',
                  {'success': 'Your timetable has been successfully generated', 'tt': tt, 'timeslots': timeslots, 'subjects': subjects})

def create_pdf(request):
    data = {
        'a': 'noura amora',
        'b': to_pdf,
        'c': 'baba samy <3 ',  # 7aga htegi mn el auth
        'd': 1233434,
    }
    pdf = render_to_pdf('services/pdf.html', data)
    return HttpResponse(pdf, content_type='application/pdf')

# startdate = datetime.strptime(startdate, "%H:%M")
# print(type(startdate))
#
# print("in H: " + str(startdate.hour))
# print("in M: " + str(startdate.minute))
#
# print(startdate)
# print(startdate + timedelta(hours=startdate.hour, minutes=startdate.minute))
#
# startdate = startdate.strftime("%H:%M")
# print(startdate)
